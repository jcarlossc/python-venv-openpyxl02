# Importa módulo Openpyxl
from openpyxl import load_workbook

# Carrega a planilha do diretório planilhas/
planilha_carregada = load_workbook("planilhas/planilha.xlsx")

# Selecionar a planilha ativa
planilha_ativa = planilha_carregada.active

# Acessando células individuais
print("\nNome:", planilha_ativa["A2"].value) 
print("Idade:", planilha_ativa["B2"].value) 
print("Peso:",planilha_ativa["C2"].value) 
print("Altura:", planilha_ativa["D2"].value, "\n") 

# Acessando por linhas e colunas
print("Nome:", planilha_ativa.cell(row=3, column=1).value)
print("Idade:", planilha_ativa.cell(row=3, column=2).value)
print("Peso:", planilha_ativa.cell(row=3, column=3).value)
print("Altura:", planilha_ativa.cell(row=3, column=4).value, "\n")

# Modificando Valores em uma Planilha (Células específicas)
planilha_ativa["B2"] = 25                               # Altera a idade de Carlos
planilha_ativa["D2"] = 1.62                             # Altera a altura de Carlos
planilha_carregada.save("planilhas/planilha.xlsx")      # Salva modificação